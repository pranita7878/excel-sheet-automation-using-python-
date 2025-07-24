import openpyxl
import pandas as pd 
wb = openpyxl.load_workbook("balance.xlsx")
#Print all sheet name 
print(wb.sheetnames)

# #print score sheet
# ws = wb['Score']
# print(ws)

#print sheet1 
ws1 = wb['Sheet1']
print(ws1)

# #How to create new excel sheet 
# wb.create_sheet('NewSheet')
# #save changes permanently 
# wb.save("balance.xlsx") 

# we can create new sheet at first position also using first indexing 
# wb.create_sheet('NewSheet_2',0)
# #save changes permanently 
# wb.save("balance.xlsx")

#How to access individual cell data 
# print(ws1['B5'].value)

# #another method to access individual cell data 
# print(ws1.cell(row=6, column=1).value)

# #How to access group of rows 
# value_range = ws1['A2': 'B5']
# for a,b in value_range:
#     print(a.value, b.value)

#How to use iterrows 
# rows = ws1.iter_rows(min_row=1,max_row=7,min_col=1,max_col=2)
# print(rows)

# names = []
# balance = []
# for a,b in rows:
#     names.append(a.value)
#     balance.append(b.value)

# print(names)
# print(balance)

#print all columns 
# columns = ws1.iter_cols(min_row=1, max_row=5, min_col=1, max_col=2)
# for col in columns:
#     print(col)

# #change single value of cell 
# ws1['B5'].value = 9 
# wb.save("balance.xlsx")

#add or change new cell with value 
# ws1['A9']= 'Rick'
# ws1['B9']= 1500
# wb.save("balance.xlsx")

#Add new column name 
# ws1['C1'] = "Double Balance"
# wb.save("balance.xlsx")

#Multiply the column b value with 2 and add result in double balance column 
# for i in range(2,10):
#     b_col = ws1.cell(row=i,column=2).value
#     print(b_col) #extract  second column data 
#     c_value = b_col * 2 
#     ws1.cell(row=i,column=3).value = c_value 

# wb.save("balance.xlsx")

#How to apply font 
from openpyxl.styles import Font,Color,PatternFill

# font_style = Font(name="Chalkboard",size=14,color="1A4FDF", italic=True,bold=True)
# a4 = ws1['A4']
# a4.font = font_style
# wb.save("balance.xlsx")

#Apply font on multiple rows 
# col_style = Font(name="Reem Kufi", size=12, color="DB3B22",
#                  underline='single', strikethrough=True)

# for i in range(2,10):
#     ws1.cell(row=i, column=3).font = col_style
# wb.save("balance.xlsx")

#Apply Pattern fill and background color 
# fill_pattern = PatternFill(patternType='solid',fgColor='C64747')
# ws1['B4'].fill = fill_pattern
# wb.save("Balance.xlsx")

#Apply border style to cell 
# from openpyxl.styles import Border,Side

# top = Side(border_style='dashed',color="FF0707")
# border = Border(top=top)
# ws1['B6'].border = border
# wb.save("balance.xlsx")

#Automate Excel Formulas 
# ws1['B11'] = "=SUM(B2:B9)"
# ws1['B12'] = "=AVERAGE(B2:B9)"
# wb.save("balance.xlsx")

#Remove Special characters and quotes 
import openpyxl 
data = openpyxl.load_workbook('sample_names.xlsx')
names = data['Sheet1']

#if data is not present it will set to false 
# is_data = True 
# row_count = 1 
# #First we have to count and fetch all the rows data
# while is_data:
#     row_count +=1 
#     first_name = names.cell(row=row_count, column=1).value
#     if first_name != None:
#        names.cell(row=row_count,column=1).value = first_name.strip('"') 
#     else:
#         is_data = False 

# data.save('sample_names.xlsx')

